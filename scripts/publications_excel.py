"""Excel-based curation console for the publication list.

Pipeline:
    Google Scholar
      -> scripts/update_publications.py             -> publications-data-raw.json  (raw snapshot, never edited)
      -> scripts/publications_excel.py refresh      -> publications-review.xlsx    (YOU edit this in Excel)
      -> scripts/publications_excel.py build        -> publications-data.json       (website data, generated)

In publications-review.xlsx:
    Publications sheet - one row per paper
        show      yes/no  -> should the paper appear on the website?
        category  dropdown of "Area / Topic" (drives the Research pages)
        year/authors/title/journal/volume/issue/pages -> fix anything Scholar got wrong
        status    NEW = appeared in the latest scrape, GONE = no longer on Scholar
        notes     free text for yourself
    Journals sheet - impact factor, JCR quartile, and whether the journal gets a
        row in the statistics table (papers stay visible either way).
    Categories sheet - the Area/Topic list, read from the R-*.html research pages.

Re-running "export" keeps everything you typed (matched by the Scholar id),
appends newly scraped papers at the top of their year and marks papers that
disappeared from Scholar with status GONE. Rows you added by hand (no id) are
kept as well, so the Excel can also hold papers Scholar does not list.

All counts on the website are recalculated by "build" - never edit
publications-data.json by hand.

Usage:
    python scripts/publications_excel.py refresh        # scrape Scholar + update the workbook
    python scripts/publications_excel.py build          # workbook -> publications-data.json
    python scripts/publications_excel.py build --check  # report only, write nothing
    python scripts/publications_excel.py export         # workbook update without scraping again
"""
import argparse
import difflib
import html as html_lib
import json
import re
import sys
from collections import Counter, OrderedDict
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from update_publications import split_venue, normalize_journal  # noqa: E402
import update_publications  # noqa: E402  (used by the "refresh" command)

RAW_FILE = ROOT / "publications-data-raw.json"
BOOK_FILE = ROOT / "publications-review.xlsx"
PUBLISHED_FILE = ROOT / "publications-data.json"
RESEARCH_PAGES = ["R-HPS.html", "R-ASD.html", "R-SF.html", "R-AI.html"]

# (column title, width, editable by you?)
COLUMNS = [
    ("id", 26, False),
    ("show", 7, True),
    ("year", 7, True),
    ("authors", 40, True),
    ("title", 62, True),
    ("journal", 36, True),
    ("volume", 8, True),
    ("issue", 7, True),
    ("pages", 12, True),
    ("category", 44, True),
    ("status", 8, False),
    ("notes", 26, True),
    ("venue_raw", 40, False),
    ("scholar_url", 34, False),
]
COL = {name: index + 1 for index, (name, _, _) in enumerate(COLUMNS)}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YES_NO = ("yes", "no")


def truthy(value) -> bool:
    return str(value).strip().lower() in ("yes", "y", "true", "1", "x")


def short(text: str, width: int = 60) -> str:
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    return text if len(text) <= width else text[: width - 3] + "..."


def plain_text(fragment: str) -> str:
    text = html_lib.unescape(re.sub(r"<[^>]+>", " ", fragment or ""))
    return re.sub(r"\s+", " ", text).strip()


def title_key(text: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(text or "").lower())).strip()


def scholar_id(item: dict) -> str:
    url = str(item.get("citationUrl") or "").replace("&amp;", "&")
    match = re.search(r"citation_for_view=([^&\s]+)", url)
    return match.group(1) if match else ""


def load_json(path: Path):
    if not path.exists() or path.stat().st_size == 0:
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"! {path.name} is not valid JSON ({exc}); ignoring it.")
        return None



class _ResearchPageParser(HTMLParser):
    """Reads a research page: h1 = area, every .box-title = one topic card."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.area = ""
        self.cards = []
        self._depth = 0
        self._box_depth = None
        self._buffer = []

    def handle_starttag(self, tag, attrs):
        if tag == "div":
            self._depth += 1
            classes = (dict(attrs).get("class") or "").split()
            if "box-title" in classes and self._box_depth is None:
                self._box_depth = self._depth
                self.cards.append({"topic": "", "entries": []})
        if tag == "p":
            self._flush_paragraph()      # some cards use <p> without </p>
        if tag in ("h1", "h2", "p"):
            self._buffer = []

    def handle_startendtag(self, tag, attrs):
        return

    def _flush_paragraph(self):
        text = re.sub(r"\s+", " ", "".join(self._buffer)).strip()
        if text and self._box_depth is not None and self.cards:
            self.cards[-1]["entries"].append(text)
        self._buffer = []

    def handle_endtag(self, tag):
        if tag in ("h1", "h2", "p"):
            text = re.sub(r"\s+", " ", "".join(self._buffer)).strip()
            if tag == "h1" and not self.area:
                self.area = text
            elif tag == "h2" and self._box_depth is not None and self.cards:
                self.cards[-1]["topic"] = text
            elif tag == "p" and self._box_depth is not None and self.cards and text:
                self.cards[-1]["entries"].append(text)
            self._buffer = []
        if tag == "div":
            if self._box_depth == self._depth:
                self._flush_paragraph()  # last <p> of the card was never closed
                self._box_depth = None
            self._depth -= 1

    def handle_data(self, data):
        self._buffer.append(data)


def read_research_pages() -> list:
    """[{page, area, topic, entries}] for every topic card of the research pages."""
    cards = []
    for page in RESEARCH_PAGES:
        path = ROOT / page
        if not path.exists():
            continue
        parser = _ResearchPageParser()
        parser.feed(path.read_text(encoding="utf-8", errors="replace"))
        for card in parser.cards:
            if not card["topic"]:
                continue
            cards.append({
                "page": page,
                "area": parser.area or path.stem,
                "topic": card["topic"],
                "entries": card["entries"],
            })
    return cards


def read_categories() -> list:
    return [{"area": card["area"], "topic": card["topic"], "page": card["page"]}
            for card in read_research_pages()]


def categories_from_pages(raw_items: list):
    """Match the papers already listed on the research pages to the scraped records.

    An article/page number (5+ digits) is a unique fingerprint, so it is tried
    first; only then a strict fuzzy title match is used.

    Returns (id -> "Area / Topic", list of page entries that could not be matched).
    """
    by_title = {title_key(item["title"]): scholar_id(item) for item in raw_items}
    by_number = {}
    for item in raw_items:
        text = f"{item.get('venue') or ''}"
        for number in re.findall(r"\d{5,}", text):
            by_number.setdefault(number, set()).add(scholar_id(item))

    lookup = {}
    unmatched = []
    for card in read_research_pages():
        label = f"{card['area']} / {card['topic']}"
        for entry in card["entries"]:
            identifier = ""
            candidates = set()
            for number in re.findall(r"\d{5,}", entry):
                found = by_number.get(number, set())
                if len(found) == 1:
                    candidates |= found
            if len(candidates) == 1:
                identifier = next(iter(candidates))
            else:
                title_match = re.search(r"\(\d{4}\)\.\s*(.+?)\.\s+[A-Z\u00c0-\u024f]", entry)
                key = title_key(title_match.group(1)) if title_match else ""
                best = difflib.get_close_matches(key, list(by_title), n=1, cutoff=0.82) if key else []
                if best:
                    identifier = by_title[best[0]]
            if identifier:
                lookup[identifier] = label
            else:
                unmatched.append((label, entry))
    return lookup, unmatched


def read_workbook_rows():
    """Everything you typed last time, keyed by Scholar id (or 'manual:<title>')."""
    if not BOOK_FILE.exists():
        return OrderedDict()
    book = load_workbook(BOOK_FILE, data_only=True)
    if "Publications" not in book.sheetnames:
        return OrderedDict()
    rows = OrderedDict()
    for values in book["Publications"].iter_rows(min_row=2, values_only=True):
        record = {name: values[index] if index < len(values) else None
                  for index, (name, _, _) in enumerate(COLUMNS)}
        if not (record.get("title") or record.get("id")):
            continue
        key = str(record.get("id") or "").strip() or "manual:" + title_key(record.get("title"))
        record["id"] = key
        rows[key] = record
    return rows


def read_journal_settings() -> dict:
    settings = {}
    if not BOOK_FILE.exists():
        return settings
    book = load_workbook(BOOK_FILE, data_only=True)
    if "Journals" not in book.sheetnames:
        return settings
    for journal, _count, impact, quartile, in_stats in book["Journals"].iter_rows(
        min_row=2, max_col=5, values_only=True
    ):
        if not journal:
            continue
        settings[str(journal).strip()] = {
            "impact": "" if impact is None else str(impact).strip(),
            "quartile": "" if quartile is None else str(quartile).strip(),
            "inStats": True if in_stats is None else truthy(in_stats),
        }
    return settings


def bootstrap_from_published(raw_items: list):
    """Migrate the decisions already visible on the website into the first workbook."""
    published = load_json(PUBLISHED_FILE) or {}
    shown = {scholar_id(item) for group in published.get("years") or [] for item in group.get("items") or []}
    stats_journals = {str(stat.get("journal")) for stat in published.get("stats") or []}
    journals = {}
    for journal in stats_journals:
        journals[journal] = {"impact": "", "quartile": "", "inStats": True}
    # Venues that are on the website but have no statistics row keep that behaviour.
    for group in published.get("years") or []:
        for item in group.get("items") or []:
            journal = str(item.get("journal") or "").strip()
            if journal:
                journals.setdefault(journal, {"impact": "", "quartile": "", "inStats": journal in stats_journals})
    for journal, impact in (published.get("impactFactors") or {}).items():
        journals.setdefault(journal, {"impact": "", "quartile": "", "inStats": False})
        journals[journal]["impact"] = str(impact)
    for journal, quartile in (published.get("quartiles") or {}).items():
        journals.setdefault(journal, {"impact": "", "quartile": "", "inStats": False})
        journals[journal]["quartile"] = str(quartile)
    return shown, journals


def merge_rows(raw_items: list, previous: dict):
    """Combine the raw scrape with what you typed before; keep manual rows.

    Returns (rows, unmatched research-page entries). Empty category cells are
    always re-filled from the research pages, so adding a paper to a topic card
    by hand (or fixing a title) is picked up on the next run.
    """
    first_run = not previous
    prefilled, unmatched = categories_from_pages(raw_items)
    shown_before, _ = bootstrap_from_published(raw_items) if first_run else (set(), {})

    merged = []
    seen = set()
    for item in raw_items:
        identifier = scholar_id(item) or "manual:" + title_key(item["title"])
        seen.add(identifier)
        journal, volume, issue, pages = split_venue(item.get("venue") or "")
        stored = previous.get(identifier, {})
        category = str(stored.get("category") or "").strip() or prefilled.get(identifier, "")
        row = {
            "id": identifier,
            "show": stored.get("show") if stored else ("yes" if (not first_run or identifier in shown_before) else "no"),
            "year": stored.get("year") or item.get("year"),
            "authors": stored.get("authors") or item.get("authors") or "",
            "title": stored.get("title") or item.get("title") or "",
            "journal": stored.get("journal") or normalize_journal(item.get("venue") or "") or journal,
            "volume": stored.get("volume") if stored.get("volume") not in (None, "") else volume,
            "issue": stored.get("issue") if stored.get("issue") not in (None, "") else issue,
            "pages": stored.get("pages") if stored.get("pages") not in (None, "") else pages,
            "category": category,
            "status": "ok" if stored else "NEW",
            "notes": stored.get("notes") or "",
            "venue_raw": item.get("venue") or "",
            "scholar_url": str(item.get("citationUrl") or "").replace("&amp;", "&"),
        }
        if row["show"] in (None, ""):
            row["show"] = "yes"
        merged.append(row)

    # Papers that are no longer on Scholar (or that you typed by hand) are kept.
    for identifier, stored in previous.items():
        if identifier in seen:
            continue
        row = {name: stored.get(name) for name, _, _ in COLUMNS}
        row["id"] = identifier
        row["status"] = "manual" if identifier.startswith("manual:") else "GONE"
        row["show"] = row.get("show") or "no"
        merged.append(row)

    merged.sort(key=lambda row: -(int(row["year"]) if str(row.get("year") or "").isdigit() else 0))
    return merged, unmatched


def write_workbook(rows: list, journal_settings: dict) -> None:
    book = Workbook()

    sheet = book.active
    sheet.title = "Publications"
    for index, (name, width, editable) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "C2"

    for row_index, row in enumerate(rows, start=2):
        for name, _, editable in COLUMNS:
            cell = sheet.cell(row=row_index, column=COL[name], value=row.get(name))
            if editable:
                cell.fill = EDIT_FILL
        sheet.cell(row=row_index, column=COL["status"]).alignment = Alignment(horizontal="center")

    last_row = max(len(rows) + 1, 2)
    categories = read_categories()
    labels = [f"{entry['area']} / {entry['topic']}" for entry in categories]

    category_sheet = book.create_sheet("Categories")
    category_sheet.append(["area", "topic", "label (paste this into the category column)", "page"])
    for index in range(1, 5):
        category_sheet.cell(row=1, column=index).fill = HEADER_FILL
        category_sheet.cell(row=1, column=index).font = HEADER_FONT
    for entry, label in zip(categories, labels):
        category_sheet.append([entry["area"], entry["topic"], label, entry["page"]])
    for column, width in zip("ABCD", (30, 46, 52, 14)):
        category_sheet.column_dimensions[column].width = width

    if labels:
        validation = DataValidation(
            type="list", formula1=f"=Categories!$C$2:$C${len(labels) + 1}", allow_blank=True
        )
        validation.error = "Pick an Area / Topic from the Categories sheet."
        sheet.add_data_validation(validation)
        column = get_column_letter(COL["category"])
        validation.add(f"{column}2:{column}{last_row}")

    show_validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
    show_validation.error = "Type yes or no."
    sheet.add_data_validation(show_validation)
    show_column = get_column_letter(COL["show"])
    show_validation.add(f"{show_column}2:{show_column}{last_row}")

    journal_sheet = book.create_sheet("Journals")
    journal_sheet.append(["journal", "papers (auto)", "impact factor", "quartile", "show in stats table"])
    for index in range(1, 6):
        journal_sheet.cell(row=1, column=index).fill = HEADER_FILL
        journal_sheet.cell(row=1, column=index).font = HEADER_FONT
    counts = Counter(str(row.get("journal") or "") for row in rows if truthy(row.get("show")))
    names = [name for name in sorted(set(list(counts) + list(journal_settings)),
                                     key=lambda n: (-counts.get(n, 0), n.lower())) if name]
    for journal in names:
        setting = journal_settings.get(journal, {})
        journal_sheet.append([
            journal,
            counts.get(journal, 0),
            setting.get("impact", ""),
            setting.get("quartile", ""),
            "yes" if setting.get("inStats", True) else "no",
        ])
    for column, width in zip("ABCDE", (58, 13, 14, 10, 20)):
        journal_sheet.column_dimensions[column].width = width
    stats_validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
    journal_sheet.add_data_validation(stats_validation)
    stats_validation.add(f"E2:E{max(len(names) + 1, 2)}")
    journal_sheet.freeze_panes = "B2"

    try:
        book.save(BOOK_FILE)
    except PermissionError:
        raise SystemExit(
            f"! Cannot write {BOOK_FILE.name} because it is open in Excel.\n"
            f"  Close the workbook and run the command again."
        )



def compose_venue(row: dict) -> str:
    """Rebuild a Scholar-like venue string from the (possibly corrected) columns."""
    journal = str(row.get("journal") or "").strip()
    volume = str(row.get("volume") or "").strip()
    issue = str(row.get("issue") or "").strip()
    pages = str(row.get("pages") or "").strip()
    year = str(row.get("year") or "").strip()
    head = journal
    if volume:
        head = f"{head} {volume}" if head else volume
        if issue:
            head = f"{head} ({issue})"
    if pages:
        head = f"{head}, {pages}" if head else pages
    return f"{head} , {year}".strip() if head or year else ""


def build_from_workbook(check_only: bool = False) -> int:
    rows = read_workbook_rows()
    if not rows:
        print(f"! {BOOK_FILE.name} not found or empty - run: python scripts/publications_excel.py export")
        return 1
    settings = read_journal_settings()

    kept = []
    problems = []
    for row in rows.values():
        if not truthy(row.get("show")):
            continue
        try:
            year = int(str(row.get("year")).strip())
        except (TypeError, ValueError):
            problems.append(f"missing/invalid year: {str(row.get('title'))[:60]}")
            year = None
        category = str(row.get("category") or "").strip()
        if category:
            area, _, topic = (part.strip() for part in category.partition("/"))
        else:
            area, topic = "", ""
        kept.append({
            "title": str(row.get("title") or "").strip(),
            "authors": str(row.get("authors") or "").strip(),
            "journal": str(row.get("journal") or "").strip(),
            "volume": str(row.get("volume") or "").strip(),
            "issue": str(row.get("issue") or "").strip(),
            "pages": str(row.get("pages") or "").strip(),
            "year": year,
            "venue": str(row.get("venue_raw") or "").strip() or compose_venue(row),
            "area": area,
            "topic": topic,
            "category": category,
            "citationUrl": str(row.get("scholar_url") or "").strip(),
        })

    kept.sort(key=lambda item: -(item["year"] or 0))

    by_year = {}
    for item in kept:
        by_year.setdefault(item["year"] or "Unknown", []).append(item)
    ordered = sorted(by_year, key=lambda y: (0, -y) if isinstance(y, int) else (1, str(y)))
    grouped = [{"year": year, "items": by_year[year]} for year in ordered]

    counts = Counter(item["journal"] for item in kept if item["journal"])
    stats = []
    for journal, count in counts.most_common():
        setting = settings.get(journal, {})
        if not setting.get("inStats", True):
            continue
        impact = str(setting.get("impact", "")).strip()
        quartile = str(setting.get("quartile", "")).strip()
        if impact.lower() in ("", "n/a", "na", "none"):
            display = ""
        elif quartile:
            display = f"(JCR {quartile}, IF={impact})"
        else:
            display = f"(IF={impact})"
        stats.append({"journal": journal, "count": count, "impactDisplay": display})

    category_counts = Counter(item["category"] for item in kept if item["category"])
    categories = [
        {"area": entry["area"], "topic": entry["topic"], "page": entry["page"],
         "count": category_counts.get(f"{entry['area']} / {entry['topic']}", 0)}
        for entry in read_categories()
    ]

    payload = {
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": BOOK_FILE.name,
        "note": "Generated by scripts/publications_excel.py build - edit publications-review.xlsx, not this file.",
        "totalCount": len(kept),
        "countedInStats": sum(stat["count"] for stat in stats),
        "stats": stats,
        "categories": categories,
        "years": grouped,
        "items": kept,
    }

    print(f"Rows in workbook      : {len(rows)}")
    print(f"Shown on the website  : {len(kept)}")
    print(f"Hidden (show = no)    : {len(rows) - len(kept)}")
    print(f"Per year              : {[(g['year'], len(g['items'])) for g in grouped]}")
    print(f"Without a category    : {sum(1 for item in kept if not item['category'])}")
    print("\nStatistics table:")
    for stat in stats:
        print(f"   {stat['journal'][:55]:55} {stat['count']:>3}  {stat['impactDisplay']}")
    print(f"   {'TOTAL counted':55} {payload['countedInStats']:>3}")
    print("\nPublications per research topic:")
    for entry in categories:
        print(f"   {entry['area'][:28]:28} {entry['topic'][:38]:38} {entry['count']:>3}")
    for problem in problems:
        print(f"! {problem}")

    if check_only:
        print("\nCheck only - publications-data.json was not touched.")
        return 0

    PUBLISHED_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {PUBLISHED_FILE.name}: {len(kept)} publications, {len(stats)} journal rows.")
    return 0



def export_workbook() -> int:
    raw = load_json(RAW_FILE)
    if not raw or not raw.get("items"):
        print(f"! {RAW_FILE.name} is missing or empty - run: python scripts/update_publications.py")
        return 1

    previous = read_workbook_rows()
    settings = read_journal_settings()
    if not settings:
        _, settings = bootstrap_from_published(raw["items"])

    rows, unmatched = merge_rows(raw["items"], previous)
    write_workbook(rows, settings)

    new_rows = [row for row in rows if row.get("status") == "NEW"]
    gone_rows = [row for row in rows if row.get("status") == "GONE"]
    shown = sum(1 for row in rows if truthy(row.get("show")))
    print(f"Wrote {BOOK_FILE.name}")
    print(f"   rows              : {len(rows)} ({shown} with show = yes)")
    print(f"   scraped           : {len(raw['items'])} records, {raw.get('generatedAt') or '?'}")
    print(f"   without category  : {sum(1 for row in rows if not str(row.get('category') or '').strip())}")
    if not previous:
        print("   first run: show / IF / quartile taken from publications-data.json,")
        print("              categories taken from the R-*.html research pages")

    assigned = Counter(str(row.get("category") or "").strip() for row in rows
                       if truthy(row.get("show")) and str(row.get("category") or "").strip())
    cards = read_research_pages()
    print(f"\nResearch page cards <-> workbook ({len(cards)} cards):")
    for card in cards:
        label = f"{card['area']} / {card['topic']}"
        state = "auto" if assigned.get(label) else "hand-written (no category assigned yet)"
        print(f"   {card['page']:<12} {short(card['topic'], 40):40} excel: {assigned.get(label, 0):>2}"
              f"  page: {len(card['entries']):>2}  -> {state}")

    if unmatched:
        print(f"\nEntries written on a page that do not match any scraped record ({len(unmatched)}):")
        for label, entry in unmatched[:20]:
            print(f"   {short(label, 40):40} {short(entry, 70)}")
        print("   -> set their category in Excel by hand (or check the title spelling)")

    print(f"\nNEW since your last review ({len(new_rows)}):")
    for row in new_rows[:40]:
        print(f"   [{row.get('year')}] {str(row.get('journal') or 'no venue')[:34]:34} {str(row.get('title'))[:58]}")
    if not new_rows:
        print("   -")
    if gone_rows:
        print(f"\nNo longer on Scholar ({len(gone_rows)}) - marked GONE in the workbook:")
        for row in gone_rows[:20]:
            print(f"   [{row.get('year')}] {str(row.get('title'))[:70]}")
    print("\nNext: set show / fix data / pick category in Excel, then run:")
    print("   python scripts/publications_excel.py build")
    return 0


def print_status() -> int:
    """Default action: is publications-data.json up to date with the workbook?"""
    raw = load_json(RAW_FILE) or {}
    published = load_json(PUBLISHED_FILE) or {}
    print("Publication pipeline")
    print(f"   1. {RAW_FILE.name:26} {len(raw.get('items') or [])} records"
          f"   scraped {raw.get('generatedAt') or '-'}")
    if BOOK_FILE.exists():
        rows = read_workbook_rows()
        shown = sum(1 for row in rows.values() if truthy(row.get("show")))
        saved = datetime.fromtimestamp(BOOK_FILE.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        print(f"   2. {BOOK_FILE.name:26} {len(rows)} rows, {shown} with show = yes   saved {saved}")
    else:
        print(f"   2. {BOOK_FILE.name:26} missing")
    print(f"   3. {PUBLISHED_FILE.name:26} {published.get('totalCount', 0)} publications"
          f"   built {published.get('generatedAt') or '-'}")

    stale = (BOOK_FILE.exists() and PUBLISHED_FILE.exists()
             and BOOK_FILE.stat().st_mtime > PUBLISHED_FILE.stat().st_mtime)
    print()
    if stale:
        print("!! The workbook is newer than the website data - your Excel edits are NOT online yet.")
        print("   Run:  python scripts/publications_excel.py build")
    else:
        print("The website data matches the workbook.")
    print("\nCommands:")
    print("   refresh          scrape Google Scholar + update the workbook")
    print("   build            workbook -> publications-data.json (run this after editing Excel)")
    print("   build --check    show what build would write, without writing")
    print("   export           update the workbook without scraping again")
    return 0


def main(argv=None) -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    parser = argparse.ArgumentParser(
        description="Excel-based curation of the publication list.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    sub = parser.add_subparsers(dest="command")
    sub.add_parser("refresh", help="scrape Google Scholar and update publications-review.xlsx (one step)")
    sub.add_parser("export", help="update publications-review.xlsx from the existing raw snapshot")
    build = sub.add_parser("build", help="turn publications-review.xlsx into publications-data.json")
    build.add_argument("--check", action="store_true", help="report only, do not write the JSON")
    args = parser.parse_args(argv)

    if args.command == "build":
        return build_from_workbook(check_only=args.check)
    if args.command == "export":
        return export_workbook()
    if args.command == "refresh":
        print("1/2  scraping Google Scholar ...")
        update_publications.main()
        print("\n2/2  updating the review workbook ...")
        return export_workbook()
    return print_status()


if __name__ == "__main__":
    raise SystemExit(main())

